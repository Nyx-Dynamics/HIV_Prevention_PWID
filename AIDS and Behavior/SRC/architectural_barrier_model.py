#!/usr/bin/env python3
"""
Architectural Barrier Model: HIV Prevention Cascade Modeling for PWID
======================================================================

Monte Carlo simulation modeling nested barriers to HIV prevention for people 
who inject drugs (PWID), demonstrating the mathematical impossibility of 
achieving R(0)=0 under current policy conditions.

PRIMARY HYPOTHESES:
1. Nested barriers create conditions where R(0)=0 is mathematically impossible
   for PWID regardless of pharmacological efficacy
2. Stochastic avoidance (probability) rather than intervention has been the
   primary mechanism preventing catastrophic HIV outbreaks among PWID
3. Introduction of methamphetamine into non-MSM PWID populations increases
   outbreak probability through network bridging and behavioral factors

THREE-LAYER BARRIER FRAMEWORK:
Layer 1: Pathogen Biology (irreversible R(0)>0 within hours)
Layer 2: HIV Testing Failures (acute infection detection gaps)
Layer 3: Architectural Barriers
    - Policy (criminalization, incarceration)
    - Stigma (healthcare discrimination, disclosure barriers)
    - Infrastructure (cascade implementation built for MSM)
    - Research Exclusion (LOOCV framework)
    - Machine Learning (algorithmic deprioritization)

Usage:
    python architectural_barrier_model.py --output-dir ../data/csv_xlsx
    python architectural_barrier_model.py --n-individuals 100000 --n-sa-sims 10000

Author: AC Demidont, DO / Nyx Dynamics LLC
Date: January 2026
Repository: https://github.com/Nyx-Dynamics/HIV_Prevention_PWID
"""

import numpy as np
import random
import logging
import json
import csv
import os
import argparse
from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Optional
from enum import Enum
from datetime import datetime
from pathlib import Path

# Optional Excel export
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Set random seeds for reproducibility
np.random.seed(42)
random.seed(42)


# =============================================================================
# ENUMERATIONS AND CONSTANTS
# =============================================================================

class BarrierLayer(Enum):
    """Three-layer barrier framework"""
    PATHOGEN_BIOLOGY = "pathogen_biology"
    HIV_TESTING = "hiv_testing"
    ARCHITECTURAL = "architectural"


class ArchitecturalSubtype(Enum):
    """Subtypes of architectural barriers"""
    POLICY = "policy"
    STIGMA = "stigma"
    INFRASTRUCTURE = "infrastructure"
    RESEARCH_EXCLUSION = "research_exclusion"
    MACHINE_LEARNING = "machine_learning"


# Literature-derived constants
LITERATURE_PARAMS = {
    # Global PWID epidemiology (Degenhardt et al., 2017)
    "global_pwid_population": 15_600_000,
    "global_pwid_hiv_prevalence": 0.178,
    "us_pwid_population": 3_500_000,

    # Methamphetamine HIV risk (Plankey 2007, Grov 2020)
    "meth_hr_hiv_seroconversion": 1.46,
    "meth_persistent_aor": 7.11,
    "meth_opioid_couse_2012": 0.043,
    "meth_opioid_couse_2018": 0.143,

    # Criminalization impact (DeBeck et al., 2017)
    "criminalization_negative_effect_rate": 0.80,
    "incarceration_hiv_rr": 1.81,

    # Healthcare stigma (Biancarelli 2019, Muncan 2020)
    "pwid_stigma_experienced": 0.88,
    "pwid_stigma_affecting_care": 0.78,
    "rural_pwid_stigma": 0.62,

    # Housing instability (Arum et al., 2021)
    "unstable_housing_hiv_arr": 1.39,
    "pwid_homelessness_rate": 0.685,

    # PrEP cascade failure (Mistler et al., 2021)
    "pwid_prep_uptake": 0.015,
    "global_pwid_art_coverage": 0.04,

    # WHO intervention requirements vs reality
    "who_required_syringes_per_pwid": 200,
    "actual_syringes_per_pwid": 22,
    "who_required_oat_coverage": 0.40,
    "actual_oat_coverage": 0.08,

    # LAI-PrEP resistance (HPTN 083, Eshleman 2022)
    "cab_la_insti_resistance_rate": 0.63,
    "cab_la_detection_delay_median_days": 98,

    # Vulnerable counties (Van Handel et al., 2016)
    "vulnerable_us_counties": 220,
    "vulnerable_counties_with_ssp": 0.21,

    # Outbreak data
    "scott_county_cases": 215,
    "massachusetts_cases": 180,
    "cabell_county_cases": 82,
    "kanawha_county_cases": 85,

    # Stochastic factors (Des Jarlais et al., 2022)
    "baseline_annual_outbreak_prob": 0.03,
    "covid_disruption_outbreak_prob": 0.25,
    "high_risk_outbreak_prob": 0.33,
}


# =============================================================================
# DATA STRUCTURES
# =============================================================================

@dataclass
class CascadeStep:
    """Single step in the PrEP cascade with barrier decomposition"""
    name: str
    description: str
    base_probability: float
    policy_penalty: float = 0.0
    stigma_penalty: float = 0.0
    infrastructure_penalty: float = 0.0
    testing_penalty: float = 0.0
    research_penalty: float = 0.0
    ml_penalty: float = 0.0


@dataclass
class PolicyScenario:
    """Policy scenario definition for counterfactual analysis"""
    name: str
    decriminalization: bool = False
    incarceration_modifier: float = 1.0
    stigma_reduction: float = 0.0
    bias_training: bool = False
    ssp_integrated: bool = False
    peer_navigation: bool = False
    low_barrier: bool = False
    trial_inclusion: bool = False
    ml_debiasing: bool = False


# =============================================================================
# CASCADE DEFINITIONS
# =============================================================================

def create_pwid_cascade() -> List[CascadeStep]:
    """Create PWID cascade with literature-derived parameters"""
    return [
        CascadeStep(
            name="awareness",
            description="Aware LAI-PrEP exists for PWID",
            base_probability=0.70,
            policy_penalty=0.25,
            stigma_penalty=0.05,
            infrastructure_penalty=0.15,
            research_penalty=0.05,
            ml_penalty=0.10,
        ),
        CascadeStep(
            name="willingness",
            description="Willing to seek despite visibility concerns",
            base_probability=0.80,
            policy_penalty=0.35,
            stigma_penalty=0.10,
            ml_penalty=0.05,
        ),
        CascadeStep(
            name="healthcare_access",
            description="Can access healthcare services",
            base_probability=0.75,
            policy_penalty=0.10,
            stigma_penalty=0.05,
            infrastructure_penalty=0.25,
        ),
        CascadeStep(
            name="disclosure",
            description="Disclose IDU to provider",
            base_probability=0.70,
            policy_penalty=0.30,
            stigma_penalty=0.15,
        ),
        CascadeStep(
            name="provider_willing",
            description="Provider willing to prescribe for PWID",
            base_probability=0.85,
            policy_penalty=0.05,
            stigma_penalty=0.25,
            research_penalty=0.10,
            ml_penalty=0.10,
        ),
        CascadeStep(
            name="hiv_testing_adequate",
            description="Adequate HIV testing before first injection",
            base_probability=0.90,
            policy_penalty=0.05,
            infrastructure_penalty=0.15,
            testing_penalty=0.25,
        ),
        CascadeStep(
            name="first_injection",
            description="Returns for first LAI injection",
            base_probability=0.75,
            policy_penalty=0.10,
            stigma_penalty=0.05,
            infrastructure_penalty=0.15,
        ),
        CascadeStep(
            name="sustained_engagement",
            description="Maintains injection schedule over time",
            base_probability=0.70,
            policy_penalty=0.20,
            stigma_penalty=0.10,
            infrastructure_penalty=0.10,
            ml_penalty=0.05,
        ),
    ]


def create_policy_scenarios() -> List[PolicyScenario]:
    """Define policy intervention scenarios"""
    return [
        PolicyScenario("Current Policy"),
        PolicyScenario(
            "Decriminalization Only",
            decriminalization=True,
            incarceration_modifier=0.3,
        ),
        PolicyScenario(
            "Decrim + Stigma Reduction",
            decriminalization=True,
            incarceration_modifier=0.3,
            stigma_reduction=0.5,
            bias_training=True,
        ),
        PolicyScenario(
            "SSP-Integrated Delivery",
            decriminalization=True,
            incarceration_modifier=0.3,
            stigma_reduction=0.7,
            bias_training=True,
            ssp_integrated=True,
            peer_navigation=True,
        ),
        PolicyScenario(
            "Full Harm Reduction",
            decriminalization=True,
            incarceration_modifier=0.0,
            stigma_reduction=0.8,
            bias_training=True,
            ssp_integrated=True,
            peer_navigation=True,
            low_barrier=True,
        ),
        PolicyScenario(
            "Full HR + PURPOSE-4 Data",
            decriminalization=True,
            incarceration_modifier=0.0,
            stigma_reduction=0.8,
            bias_training=True,
            ssp_integrated=True,
            peer_navigation=True,
            low_barrier=True,
            trial_inclusion=True,
        ),
        PolicyScenario(
            "Full HR + Algorithmic Debiasing",
            decriminalization=True,
            incarceration_modifier=0.0,
            stigma_reduction=0.8,
            bias_training=True,
            ssp_integrated=True,
            peer_navigation=True,
            low_barrier=True,
            trial_inclusion=True,
            ml_debiasing=True,
        ),
        PolicyScenario(
            "Theoretical Maximum",
            decriminalization=True,
            incarceration_modifier=0.0,
            stigma_reduction=1.0,
            bias_training=True,
            ssp_integrated=True,
            peer_navigation=True,
            low_barrier=True,
            trial_inclusion=True,
            ml_debiasing=True,
        ),
    ]


# =============================================================================
# MAIN MODEL
# =============================================================================

class ArchitecturalBarrierModel:
    """
    Monte Carlo simulation of PWID HIV prevention cascade
    with three-layer barrier decomposition.
    """

    def __init__(self, cascade: List[CascadeStep] = None):
        self.cascade = cascade or create_pwid_cascade()
        self.annual_incarceration_rate = 0.30

    def calculate_step_probability(
        self,
        step: CascadeStep,
        scenario: PolicyScenario
    ) -> Tuple[float, Dict[str, float]]:
        """
        Calculate effective step probability under given scenario.
        Returns (probability, barrier_attribution_dict)
        """
        prob = step.base_probability
        attribution = {}

        # Policy penalty
        if step.policy_penalty > 0:
            policy_effect = step.policy_penalty if not scenario.decriminalization else 0.0
            prob -= policy_effect
            attribution["policy"] = policy_effect

        # Stigma penalty
        if step.stigma_penalty > 0:
            stigma_effect = step.stigma_penalty * (1 - scenario.stigma_reduction)
            if scenario.bias_training:
                stigma_effect *= 0.7
            prob -= stigma_effect
            attribution["stigma"] = stigma_effect

        # Infrastructure penalty
        if step.infrastructure_penalty > 0:
            infra_effect = step.infrastructure_penalty
            if scenario.ssp_integrated:
                infra_effect *= 0.3
            if scenario.low_barrier:
                infra_effect *= 0.5
            prob -= infra_effect
            attribution["infrastructure"] = infra_effect

        # Testing penalty
        if step.testing_penalty > 0:
            testing_effect = step.testing_penalty
            if scenario.low_barrier:
                testing_effect *= 0.8
            prob -= testing_effect
            attribution["hiv_testing"] = testing_effect

        # Research exclusion penalty
        if step.research_penalty > 0:
            research_effect = step.research_penalty if not scenario.trial_inclusion else 0.0
            prob -= research_effect
            attribution["research_exclusion"] = research_effect

        # ML penalty
        if step.ml_penalty > 0:
            ml_effect = step.ml_penalty if not scenario.ml_debiasing else 0.0
            prob -= ml_effect
            attribution["machine_learning"] = ml_effect

        # Positive modifiers
        if scenario.ssp_integrated and step.name in ["awareness", "healthcare_access", "first_injection"]:
            prob += 0.15

        if scenario.peer_navigation and step.name in ["willingness", "first_injection", "sustained_engagement"]:
            prob += 0.10

        # Clamp to valid probability range
        prob = max(0.01, min(0.99, prob))

        return prob, attribution

    def simulate_individual(
        self,
        scenario: PolicyScenario,
        years: int = 5
    ) -> Dict:
        """Simulate one individual through the cascade"""
        result = {
            "completed_cascade": False,
            "achieved_r0_zero": False,
            "incarceration_disrupted": False,
            "failed_step": None,
            "barrier_attribution": {},
        }

        # Progress through cascade
        for step in self.cascade:
            prob, attribution = self.calculate_step_probability(step, scenario)

            if random.random() > prob:
                result["failed_step"] = step.name
                result["barrier_attribution"] = attribution
                return result

        result["completed_cascade"] = True

        # Check incarceration disruption
        annual_rate = self.annual_incarceration_rate * scenario.incarceration_modifier
        survival_prob = (1 - annual_rate) ** years

        if random.random() < survival_prob:
            result["achieved_r0_zero"] = True
        else:
            result["incarceration_disrupted"] = True

        return result

    def run_simulation(
        self,
        scenario: PolicyScenario,
        n_individuals: int = 100000,
        years: int = 5
    ) -> Dict:
        """Run full Monte Carlo simulation for a scenario"""

        results = {
            "scenario": scenario.name,
            "n_individuals": n_individuals,
            "years": years,
            "achieved_r0_zero": 0,
            "completed_cascade": 0,
            "incarceration_disrupted": 0,
            "step_failure_counts": {step.name: 0 for step in self.cascade},
            "barrier_attribution_totals": {
                "pathogen_biology": 0.0,
                "hiv_testing": 0.0,
                "policy": 0.0,
                "stigma": 0.0,
                "infrastructure": 0.0,
                "research_exclusion": 0.0,
                "machine_learning": 0.0,
            },
        }

        # Calculate step probabilities for decomposition
        step_probs = {}
        for step in self.cascade:
            prob, attr = self.calculate_step_probability(step, scenario)
            step_probs[step.name] = prob
            for barrier, impact in attr.items():
                results["barrier_attribution_totals"][barrier] += impact

        results["step_probabilities"] = step_probs

        # Theoretical calculations
        theoretical_cascade = np.prod(list(step_probs.values()))
        annual_rate = self.annual_incarceration_rate * scenario.incarceration_modifier
        incarc_survival = (1 - annual_rate) ** years
        theoretical_r0_zero = theoretical_cascade * incarc_survival

        results["theoretical_cascade_probability"] = theoretical_cascade
        results["incarceration_survival_probability"] = incarc_survival
        results["theoretical_r0_zero_probability"] = theoretical_r0_zero

        # Monte Carlo simulation
        for _ in range(n_individuals):
            individual = self.simulate_individual(scenario, years)

            if individual["achieved_r0_zero"]:
                results["achieved_r0_zero"] += 1
            if individual["completed_cascade"]:
                results["completed_cascade"] += 1
            if individual["incarceration_disrupted"]:
                results["incarceration_disrupted"] += 1
            if individual["failed_step"]:
                results["step_failure_counts"][individual["failed_step"]] += 1

        # Calculate observed rates
        results["observed_r0_zero_rate"] = results["achieved_r0_zero"] / n_individuals
        results["observed_cascade_completion_rate"] = results["completed_cascade"] / n_individuals

        # 95% CI
        p = results["observed_r0_zero_rate"]
        n = n_individuals
        se = np.sqrt(p * (1 - p) / n)
        results["r0_zero_95ci"] = (
            max(0, p - 1.96 * se),
            min(1, p + 1.96 * se)
        )

        # Barrier decomposition percentages
        total_barrier = sum(results["barrier_attribution_totals"].values())
        if total_barrier > 0:
            results["barrier_decomposition_pct"] = {
                barrier: (impact / total_barrier) * 100
                for barrier, impact in results["barrier_attribution_totals"].items()
            }

            # Three-layer grouping
            results["three_layer_decomposition"] = {
                "pathogen_biology": results["barrier_decomposition_pct"]["pathogen_biology"],
                "hiv_testing": results["barrier_decomposition_pct"]["hiv_testing"],
                "architectural": (
                    results["barrier_decomposition_pct"]["policy"] +
                    results["barrier_decomposition_pct"]["stigma"] +
                    results["barrier_decomposition_pct"]["infrastructure"] +
                    results["barrier_decomposition_pct"]["research_exclusion"] +
                    results["barrier_decomposition_pct"]["machine_learning"]
                )
            }

        return results


# =============================================================================
# MSM COMPARISON
# =============================================================================

def calculate_msm_cascade_completion() -> Dict:
    """Calculate MSM cascade for comparison (minimal barriers)"""
    msm_steps = {
        "awareness": 0.90,
        "willingness": 0.85,
        "healthcare_access": 0.80,
        "disclosure": 0.75,
        "provider_willing": 0.90,
        "hiv_testing_adequate": 0.85,
        "first_injection": 0.80,
        "sustained_engagement": 0.75,
    }

    cascade_completion = np.prod(list(msm_steps.values()))
    annual_incarc = 0.05  # 5% for MSM
    incarc_survival = (1 - annual_incarc) ** 5
    p_r0_zero = cascade_completion * incarc_survival

    return {
        "population": "MSM",
        "cascade_steps": msm_steps,
        "cascade_completion": cascade_completion,
        "incarceration_survival": incarc_survival,
        "p_r0_zero": p_r0_zero,
        "snr": 9180.0,
        "in_training_set": True,
    }


# =============================================================================
# STOCHASTIC AVOIDANCE MODEL
# =============================================================================

class StochasticAvoidanceModel:
    """
    Models the probability of catastrophic outbreak as stochastic
    avoidance fails due to increasing network density.
    """

    def __init__(self):
        self.baseline_network_density = 0.15
        self.meth_network_multiplier = 2.5
        self.critical_threshold = 0.35
        self.baseline_outbreak_prob = 0.03
        self.housing_instability = LITERATURE_PARAMS["pwid_homelessness_rate"]
        self.meth_prevalence = LITERATURE_PARAMS["meth_opioid_couse_2018"]
        self.meth_growth_rate = 0.025
        self.ssp_coverage = LITERATURE_PARAMS["vulnerable_counties_with_ssp"]
        self.oat_coverage = LITERATURE_PARAMS["actual_oat_coverage"]

    def calculate_network_density(self, year: int, base_year: int = 2024) -> float:
        """Calculate effective network density at given year"""
        years_elapsed = year - base_year

        # Methamphetamine prevalence growth
        current_meth = self.meth_prevalence * (1 + self.meth_growth_rate) ** years_elapsed
        current_meth = min(current_meth, 0.5)

        # Density components
        meth_effect = current_meth * self.meth_network_multiplier
        housing_effect = self.housing_instability * 0.5

        density = self.baseline_network_density + meth_effect + housing_effect
        return min(density, 1.0)

    def calculate_outbreak_probability(
        self,
        density: float,
        ssp_coverage: float = None,
        oat_coverage: float = None
    ) -> float:
        """Calculate annual outbreak probability given network density"""
        ssp = ssp_coverage or self.ssp_coverage
        oat = oat_coverage or self.oat_coverage

        p_outbreak = self.baseline_outbreak_prob

        # Exponential increase above critical threshold
        if density > self.critical_threshold:
            excess = density - self.critical_threshold
            p_outbreak *= np.exp(3 * excess)

        # Protective effects
        ssp_protection = 1 - (ssp * 0.4)
        oat_protection = 1 - (oat * 0.3)
        p_outbreak *= ssp_protection * oat_protection

        return min(p_outbreak, 1.0)

    def simulate_time_to_outbreak(
        self,
        n_simulations: int = 10000,
        max_years: int = 20
    ) -> Dict:
        """Simulate time until major outbreak"""
        outbreak_years = []
        no_outbreak_count = 0

        for _ in range(n_simulations):
            outbreak_occurred = False

            for year_offset in range(max_years):
                year = 2024 + year_offset
                density = self.calculate_network_density(year)
                p_outbreak = self.calculate_outbreak_probability(density)

                if random.random() < p_outbreak:
                    outbreak_years.append(year_offset)
                    outbreak_occurred = True
                    break

            if not outbreak_occurred:
                no_outbreak_count += 1

        # Calculate statistics
        if outbreak_years:
            outbreak_years = np.array(outbreak_years)
            results = {
                "median_years_to_outbreak": float(np.median(outbreak_years)),
                "mean_years_to_outbreak": float(np.mean(outbreak_years)),
                "std_years": float(np.std(outbreak_years)),
                "p10_years": float(np.percentile(outbreak_years, 10)),
                "p25_years": float(np.percentile(outbreak_years, 25)),
                "p75_years": float(np.percentile(outbreak_years, 75)),
                "p90_years": float(np.percentile(outbreak_years, 90)),
                "probability_outbreak_5_years": float(np.mean(outbreak_years <= 5)),
                "probability_outbreak_10_years": float(np.mean(outbreak_years <= 10)),
                "probability_no_outbreak": no_outbreak_count / n_simulations,
            }
        else:
            results = {
                "median_years_to_outbreak": None,
                "probability_no_outbreak": 1.0
            }

        # Add trajectory
        trajectory = []
        for year_offset in range(max_years):
            year = 2024 + year_offset
            density = self.calculate_network_density(year)
            p_outbreak = self.calculate_outbreak_probability(density)
            trajectory.append({
                "year": year,
                "network_density": density,
                "annual_outbreak_probability": p_outbreak,
            })
        results["trajectory"] = trajectory

        return results


# =============================================================================
# OUTPUT FUNCTIONS
# =============================================================================

def save_results_json(output: Dict, output_dir: str, filename: str = "architectural_barrier_results.json"):
    """Save results to JSON"""
    path = os.path.join(output_dir, filename)
    with open(path, 'w') as f:
        json.dump(output, f, indent=2, default=str)
    logger.info(f"Saved: {path}")


def save_results_csv(all_results: List[Dict], msm_results: Dict, output_dir: str, 
                     filename: str = "architectural_barrier_results.csv"):
    """Save results to CSV"""
    path = os.path.join(output_dir, filename)
    
    with open(path, 'w', newline='') as f:
        writer = csv.writer(f)
        
        # Header
        writer.writerow([
            "Scenario", "Achieved R0=0", "Completed Cascade", "Incarceration Disrupted",
            "Observed R0=0 Rate", "Observed Cascade Rate", "95% CI Lower", "95% CI Upper"
        ])
        
        # Data rows
        for r in all_results:
            writer.writerow([
                r["scenario"],
                r["achieved_r0_zero"],
                r["completed_cascade"],
                r["incarceration_disrupted"],
                f"{r['observed_r0_zero_rate']:.6f}",
                f"{r['observed_cascade_completion_rate']:.6f}",
                f"{r['r0_zero_95ci'][0]:.6f}",
                f"{r['r0_zero_95ci'][1]:.6f}"
            ])
        
        # MSM comparison
        writer.writerow([])
        writer.writerow(["MSM Comparison", "", "", "", 
                        f"{msm_results['p_r0_zero']:.6f}",
                        f"{msm_results['cascade_completion']:.6f}", "N/A", "N/A"])
        
        # Barrier decomposition
        writer.writerow([])
        writer.writerow(["Barrier Decomposition (Current Policy)", "Percentage (%)"])
        current = all_results[0]
        if "three_layer_decomposition" in current:
            for layer, pct in current["three_layer_decomposition"].items():
                writer.writerow([layer.replace('_', ' ').title(), f"{pct:.2f}"])
        
        writer.writerow([])
        writer.writerow(["Architectural Subtypes", "Percentage (%)"])
        if "barrier_decomposition_pct" in current:
            for barrier in ["policy", "stigma", "infrastructure", "research_exclusion", "machine_learning"]:
                pct = current["barrier_decomposition_pct"].get(barrier, 0)
                writer.writerow([barrier.replace('_', ' ').title(), f"{pct:.2f}"])
    
    logger.info(f"Saved: {path}")


def save_results_xlsx(all_results: List[Dict], msm_results: Dict, sa_results: Dict,
                      output_dir: str, filename: str = "architectural_barrier_results.xlsx"):
    """Save results to Excel (requires openpyxl)"""
    if not HAS_OPENPYXL:
        logger.warning("openpyxl not installed - skipping Excel export")
        return
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    wb = Workbook()
    
    # Sheet 1: Policy Scenarios
    ws1 = wb.active
    ws1.title = "Policy Scenarios"
    
    headers = ["Scenario", "P(R0=0)", "95% CI Lower", "95% CI Upper", 
               "Cascade Completion", "Incarceration Survival"]
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    
    for row, r in enumerate(all_results, 2):
        ws1.cell(row=row, column=1, value=r["scenario"])
        ws1.cell(row=row, column=2, value=r["observed_r0_zero_rate"])
        ws1.cell(row=row, column=3, value=r["r0_zero_95ci"][0])
        ws1.cell(row=row, column=4, value=r["r0_zero_95ci"][1])
        ws1.cell(row=row, column=5, value=r["observed_cascade_completion_rate"])
        ws1.cell(row=row, column=6, value=r["incarceration_survival_probability"])
    
    # MSM row
    row = len(all_results) + 2
    ws1.cell(row=row, column=1, value="MSM (Comparison)")
    ws1.cell(row=row, column=2, value=msm_results["p_r0_zero"])
    ws1.cell(row=row, column=5, value=msm_results["cascade_completion"])
    ws1.cell(row=row, column=6, value=msm_results["incarceration_survival"])
    
    # Sheet 2: Barrier Decomposition
    ws2 = wb.create_sheet("Barrier Decomposition")
    current = all_results[0]
    
    ws2.cell(row=1, column=1, value="Three-Layer Decomposition").font = Font(bold=True)
    row = 2
    if "three_layer_decomposition" in current:
        for layer, pct in current["three_layer_decomposition"].items():
            ws2.cell(row=row, column=1, value=layer.replace('_', ' ').title())
            ws2.cell(row=row, column=2, value=pct)
            row += 1
    
    row += 1
    ws2.cell(row=row, column=1, value="Architectural Subtypes").font = Font(bold=True)
    row += 1
    if "barrier_decomposition_pct" in current:
        for barrier in ["policy", "stigma", "infrastructure", "research_exclusion", "machine_learning"]:
            pct = current["barrier_decomposition_pct"].get(barrier, 0)
            ws2.cell(row=row, column=1, value=barrier.replace('_', ' ').title())
            ws2.cell(row=row, column=2, value=pct)
            row += 1
    
    # Sheet 3: Stochastic Avoidance
    ws3 = wb.create_sheet("Stochastic Avoidance")
    ws3.cell(row=1, column=1, value="Metric").font = Font(bold=True)
    ws3.cell(row=1, column=2, value="Value").font = Font(bold=True)
    
    sa_metrics = [
        ("Median Years to Outbreak", sa_results.get("median_years_to_outbreak")),
        ("Mean Years to Outbreak", sa_results.get("mean_years_to_outbreak")),
        ("P(Outbreak within 5 years)", sa_results.get("probability_outbreak_5_years")),
        ("P(Outbreak within 10 years)", sa_results.get("probability_outbreak_10_years")),
        ("P(No Outbreak 20 years)", sa_results.get("probability_no_outbreak")),
    ]
    
    for row, (metric, value) in enumerate(sa_metrics, 2):
        ws3.cell(row=row, column=1, value=metric)
        ws3.cell(row=row, column=2, value=value)
    
    # Save
    path = os.path.join(output_dir, filename)
    wb.save(path)
    logger.info(f"Saved: {path}")


# =============================================================================
# MAIN
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Architectural Barrier Model for PWID HIV Prevention",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
    python architectural_barrier_model.py
    python architectural_barrier_model.py --output-dir ../data/csv_xlsx
    python architectural_barrier_model.py --n-individuals 100000 --n-sa-sims 10000
        """
    )
    parser.add_argument("--output-dir", type=str, default="../data/csv_xlsx",
                       help="Directory for output files")
    parser.add_argument("--n-individuals", type=int, default=100000,
                       help="Individuals per scenario (default: 100000)")
    parser.add_argument("--n-sa-sims", type=int, default=10000,
                       help="Stochastic avoidance simulations (default: 10000)")
    args = parser.parse_args()
    
    # Create output directory
    os.makedirs(args.output_dir, exist_ok=True)
    
    print("=" * 80)
    print("ARCHITECTURAL BARRIER MODEL: HIV PREVENTION CASCADE MODELING")
    print("Monte Carlo Simulation with 3-Layer Barrier Framework")
    print(f"Output directory: {os.path.abspath(args.output_dir)}")
    print("=" * 80)
    print()
    
    # Initialize
    model = ArchitecturalBarrierModel()
    scenarios = create_policy_scenarios()
    
    # Run cascade simulations
    print(f"Running cascade simulations ({args.n_individuals:,} individuals per scenario)...")
    all_results = []
    for scenario in scenarios:
        print(f"  Simulating: {scenario.name}...")
        results = model.run_simulation(scenario, n_individuals=args.n_individuals, years=5)
        all_results.append(results)
    
    # MSM comparison
    print("\nCalculating MSM comparison...")
    msm_results = calculate_msm_cascade_completion()
    
    # Print results
    print("\n" + "=" * 100)
    print("RESULTS: LAI-PrEP CASCADE COMPLETION BY POLICY SCENARIO")
    print("=" * 100)
    print()
    print(f"{'Scenario':<40} {'P(R(0)=0)':<12} {'95% CI':<25} {'Cascade':<10}")
    print("-" * 90)
    
    for r in all_results:
        ci_str = f"({r['r0_zero_95ci'][0]:.4f}, {r['r0_zero_95ci'][1]:.4f})"
        print(f"{r['scenario']:<40} {r['observed_r0_zero_rate']:.4f}       {ci_str:<25} {r['observed_cascade_completion_rate']:.4f}")
    
    print("-" * 90)
    print(f"{'MSM (comparison)':<40} {msm_results['p_r0_zero']:.4f}       {'N/A':<25} {msm_results['cascade_completion']:.4f}")
    
    # Disparity
    current_pwid = all_results[0]['observed_r0_zero_rate']
    disparity = msm_results['p_r0_zero'] / current_pwid if current_pwid > 0 else float('inf')
    print(f"\nDISPARITY: MSM vs PWID (Current Policy) = {disparity:,.0f}-fold")
    
    # Barrier decomposition
    print("\n" + "=" * 80)
    print("THREE-LAYER BARRIER DECOMPOSITION (Current Policy)")
    print("=" * 80)
    current = all_results[0]
    if "three_layer_decomposition" in current:
        for layer, pct in current["three_layer_decomposition"].items():
            print(f"  {layer.replace('_', ' ').title():<30}: {pct:>6.1f}%")
    
    print("\nArchitectural Barrier Subtypes:")
    if "barrier_decomposition_pct" in current:
        for barrier in ["policy", "stigma", "infrastructure", "research_exclusion", "machine_learning"]:
            pct = current["barrier_decomposition_pct"].get(barrier, 0)
            print(f"    {barrier.replace('_', ' ').title():<28}: {pct:>6.1f}%")
    
    # Stochastic avoidance
    print("\n" + "=" * 80)
    print("STOCHASTIC AVOIDANCE FAILURE PREDICTION")
    print("=" * 80)
    
    sa_model = StochasticAvoidanceModel()
    sa_results = sa_model.simulate_time_to_outbreak(n_simulations=args.n_sa_sims)
    
    if sa_results["median_years_to_outbreak"]:
        print(f"Median years to major outbreak: {sa_results['median_years_to_outbreak']:.1f}")
        print(f"P(outbreak within 5 years):     {sa_results['probability_outbreak_5_years'] * 100:.1f}%")
        print(f"P(outbreak within 10 years):    {sa_results['probability_outbreak_10_years'] * 100:.1f}%")
    
    # Save results
    print("\n" + "=" * 80)
    print("SAVING RESULTS")
    print("=" * 80)
    
    output = {
        "timestamp": datetime.now().isoformat(),
        "model_parameters": {"literature_params": LITERATURE_PARAMS},
        "cascade_results": [{**r, "r0_zero_95ci": list(r["r0_zero_95ci"])} for r in all_results],
        "msm_comparison": msm_results,
        "disparity_fold": disparity,
        "stochastic_avoidance": {k: v for k, v in sa_results.items() if k != "trajectory"},
    }
    
    save_results_json(output, args.output_dir)
    save_results_csv(all_results, msm_results, args.output_dir)
    save_results_xlsx(all_results, msm_results, sa_results, args.output_dir)
    
    print("\nDone!")
    return all_results, sa_results


if __name__ == "__main__":
    results, sa_results = main()
